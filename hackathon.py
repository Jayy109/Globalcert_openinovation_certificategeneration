import cv2 as cv
import openpyxl

# template1.png is the template
# certificate Templates ('https://siddharthdhara17.github.io/hack/New%20folder/hack.html')

template_path = '/Users/jayz/Downloads/template.jpeg'
#Excel file containing names of
# the participants
print("Enter Path Of ExcelSheet")
details_path = '/Users/jayz/Downloads/sample2.xlsx'

# Output Paths
output_path = '/Users/jayz/Desktop'

# Setting the font size and font
# colour

font_size = 2
font_color = (0, 0, 0)

# Coordinates on the certificate where
# will be printing the name (set
# according to your own template)
coordinate_y_adjustment = 80
coordinate_x_adjustment = 2

coordinate_y_adjustment2 = -90
coordinate_x_adjustment2 = -30

coordinate_y_adjustment3 = -230
coordinate_x_adjustment3 = 2
# loading the details.xlsx workbook
# and grabbing the active sheet
obj = openpyxl.load_workbook(details_path)
sheet = obj.active

# printing for the first 10 names in the
# excel sheet
for i in range(1, 6):
    # grabs the row=i and column=1 cell
    # that contains the name value of that
    # cell is stored in the variable certi_name
    get_name = sheet.cell(row=i, column=1)
    certi_name = get_name.value

    get_project = sheet.cell(row=i, column=2)
    project_name = get_project.value

    get_pos = sheet.cell(row=i, column=3)
    pos_name = get_pos.value
    # read the certificate template
    img = cv.imread(template_path)

    # choose the font from opencv
    font = cv.FONT_HERSHEY_SCRIPT_COMPLEX
    font2 = cv.FONT_HERSHEY_SIMPLEX
    font3 = cv.FONT_HERSHEY_SIMPLEX

    # get the size of the name to be
    # printed
    text_size = cv.getTextSize(certi_name, font, font_size,1)[0]
    text2_size = cv.getTextSize(project_name, font2, font_size,1)[0]
    text3_size = cv.getTextSize(pos_name, font3, font_size, 1)[0]

    # get the (x,y) coordinates where the
    # name is to written on the template
    # The function cv.putText accepts only
    # integer arguments so convert it into 'int'.
    text_x = (img.shape[1] - text_size[0]) / 2 + coordinate_x_adjustment
    text_y = (img.shape[0] + text_size[1]) / 2 - coordinate_y_adjustment
    text_x = int(text_x)
    text_y = int(text_y)
    cv.putText(img, certi_name,
               (text_x, text_y),
               font,
               font_size,
               font_color, 3)

    text2_x = (img.shape[1] - text2_size[0]) / 2 + coordinate_x_adjustment2
    text2_y = (img.shape[0] + text2_size[1]) / 2 - coordinate_y_adjustment2
    text2_x = int(text2_x)
    text2_y = int(text2_y)
    cv.putText(img, project_name,
               (text2_x, text2_y),
               font2,
               font_size,
               font_color, 3)

    text3_x = (img.shape[1] - text3_size[0]) / 2 + coordinate_x_adjustment3
    text3_y = (img.shape[0] + text3_size[1]) / 2 - coordinate_y_adjustment3
    text3_x = int(text3_x)
    text3_y = int(text3_y)
    cv.putText(img, pos_name,
               (text3_x, text3_y),
               font3,
               font_size,
               font_color, 3)

    # Output path along with the name of the
    # certificate generated
    certi_path = output_path +'Certificate' +str(i)+ '.png'
    # Save the certificate
    cv.imwrite(certi_path, img)